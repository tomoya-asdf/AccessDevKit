import pytest
from unittest.mock import patch, MagicMock
from src.core.access_handler import access_application, export_objects
from src.core.db_operations import db_connection, get_table_names, get_table_data, run_benchmark
from src.core.reporting import create_diff_report
import os
import openpyxl

# access_application をモック化する例
def test_access_application_mock():
    with patch('src.core.access_handler.win32com.client.Dispatch') as mock_dispatch:
        mock_app = MagicMock()
        mock_dispatch.return_value = mock_app

        with access_application("dummy.accdb") as app:
            assert app == mock_app
            # モックされたオブジェクトのメソッドが呼ばれたか確認
            # withブロック終了時にQuitが呼ばれるか確認
            mock_app.Quit.assert_called_once()

# export_objects のモック化の例
def test_export_objects_mock(tmp_path):
    mock_app = MagicMock()
    # モックオブジェクトに、エクスポート対象となるオブジェクトのリストを設定
    mock_app.CurrentProject.AllForms = [MagicMock(Name="Form1"), MagicMock(Name="Form2")]
    mock_app.CurrentProject.AllReports = []
    mock_app.CurrentProject.AllModules = [MagicMock(Name="Module1")]
    mock_app.CurrentProject.AllMacros = []
    mock_app.CurrentProject.AllQueries = []
    mock_app.CurrentProject.AllTables = []

    # ExportAsText メソッドのモック
    # 実際にはファイルを作成するが、ここでは何もしないように設定
    mock_app.ExportAsText.return_value = None

    # TableDefs のモック
    mock_tdf1 = MagicMock(Name="Table1")
    mock_tdf1.Attributes = 0 # Not system or linked
    mock_tdf2 = MagicMock(Name="Table2")
    mock_tdf2.Attributes = 0 # Not system or linked
    mock_app.CurrentDb.return_value.TableDefs = [mock_tdf1, mock_tdf2]

    # DoCmd のモック
    mock_app.DoCmd = MagicMock()

    output_dir = tmp_path / "export_dir"
    output_dir.mkdir()

    exported = export_objects(mock_app, str(output_dir))

    # 期待されるカテゴリとファイル名が含まれているか確認
    assert "Forms" in exported
    assert "Form1.frm" in exported["Forms"]
    assert "Form2.frm" in exported["Forms"]
    assert "Modules" in exported
    assert "Module1.bas" in exported["Modules"]
    assert "Reports" not in exported # レポートがないことを確認
    assert "Tables" in exported
    assert "Table1.csv" in exported["Tables"]
    assert "Table2.csv" in exported["Tables"]

    # ExportAsText が適切な引数で呼び出されたか確認
    mock_app.ExportAsText.assert_any_call(0, "Form1", os.path.join(str(output_dir), "Form1.frm"))
    mock_app.ExportAsText.assert_any_call(0, "Form2", os.path.join(str(output_dir), "Form2.frm"))
    mock_app.ExportAsText.assert_any_call(5, "Module1", os.path.join(str(output_dir), "Module1.bas"))

    # DoCmd.TransferText が適切な引数で呼び出されたか確認
    mock_app.DoCmd.TransferText.assert_any_call(
        win32com.client.constants.acExportDelim,
        None,
        "Table1",
        os.path.join(str(output_dir), "Table1.csv"),
        True
    )
    mock_app.DoCmd.TransferText.assert_any_call(
        win32com.client.constants.acExportDelim,
        None,
        "Table2",
        os.path.join(str(output_dir), "Table2.csv"),
        True
    )

# import_objects のモック化の例
def test_import_objects_mock(tmp_path):
    mock_app = MagicMock()
    mock_app.LoadFromText.return_value = None
    mock_app.DoCmd = MagicMock()

    input_dir = tmp_path / "import_dir"
    input_dir.mkdir()

    # ダミーのCSVファイルを作成
    csv_file_path = input_dir / "TestTable.csv"
    csv_file_path.write_text("Header1,Header2\nValue1,Value2")

    # ダミーのfrmファイルを作成
    frm_file_path = input_dir / "TestForm.frm"
    frm_file_path.write_text("Form content")

    imported = import_objects(mock_app, str(input_dir))

    assert "Tables" in imported
    assert "TestTable.csv" in imported["Tables"]
    assert "Forms" in imported
    assert "TestForm.frm" in imported["Forms"]

    mock_app.DoCmd.TransferText.assert_any_call(
        win32com.client.constants.acImportDelim,
        None,
        "TestTable",
        str(csv_file_path),
        True
    )
    mock_app.LoadFromText.assert_any_call(0, "TestForm", str(frm_file_path))

# db_connection のモック化の例
def test_db_connection_mock():
    with patch('src.core.db_operations.pyodbc.connect') as mock_connect:
        mock_conn = MagicMock()
        mock_connect.return_value = mock_conn

        with db_connection("dummy.accdb") as conn:
            assert conn == mock_conn
            mock_conn.close.assert_called_once() # withブロック終了時にcloseが呼ばれるか確認

# get_table_names のモック化の例
def test_get_table_names_mock():
    mock_conn = MagicMock()
    mock_cursor = MagicMock()
    mock_conn.cursor.return_value = mock_cursor

    # tables() メソッドが返す値をモック
    mock_cursor.tables.return_value = [MagicMock(table_name="Table1"), MagicMock(table_name="Table2")]

    table_names = get_table_names(mock_conn)
    assert table_names == ["Table1", "Table2"]
    mock_cursor.tables.assert_called_once_with(tableType='TABLE')

# get_table_data のモック化の例
def test_get_table_data_mock():
    mock_conn = MagicMock()
    mock_cursor = MagicMock()
    mock_conn.cursor.return_value = mock_cursor

    # fetchall() メソッドが返す値をモック
    mock_cursor.fetchall.return_value = [("Data1", 1), ("Data2", 2)]

    table_data = get_table_data(mock_conn, "TestTable")
    assert table_data == {("Data1", 1), ("Data2", 2)}
    mock_cursor.execute.assert_called_once_with("SELECT * FROM [TestTable]")

# run_benchmark のモック化の例
def test_run_benchmark_mock():
    mock_conn = MagicMock()
    mock_cursor = MagicMock()
    mock_conn.cursor.return_value = mock_cursor

    # time.perf_counter をモックして、常に同じ時間を返すようにする
    with patch('src.core.db_operations.time.perf_counter', side_effect=[0, 1, 1, 2, 2, 3, 3, 4, 4, 5]): # 実行ごとに1秒かかるように設定
        timings = run_benchmark(mock_conn, "TestQuery", 5)
        assert timings == [1.0, 1.0, 1.0, 1.0, 1.0]
        assert mock_cursor.execute.call_count == 5
        assert mock_cursor.fetchall.call_count == 5

# create_diff_report のテスト
def test_create_diff_report(tmp_path):
    output_file = tmp_path / "report.xlsx"
    table_diffs = {
        "TableA": ({"row1_removed"}, set()),
        "TableB": (set(), {"row1_added"})
    }
    vba_diffs = {
        "Module1.bas": ["--- a/Module1.bas", "+++ b/Module1.bas", "@@ -1,2 +1,2 @@", "-Old line", "+New line"]
    }

    create_diff_report(table_diffs, vba_diffs, str(output_file))

    assert os.path.exists(output_file)
    # さらに詳細な内容検証を行う場合は、openpyxlでファイルを開いてシート名やセル内容を確認する
    wb = openpyxl.load_workbook(output_file)
    assert "Table_Diffs" in wb.sheetnames
    assert "VBA_Diffs" in wb.sheetnames

    ws_table = wb["Table_Diffs"]
    assert ws_table['A1'].value == "Type"
    assert ws_table['B1'].value == "Table"
    assert ws_table['A2'].value == "REMOVED"
    assert ws_table['B2'].value == "TableA"
    assert ws_table['A3'].value == "ADDED"
    assert ws_table['B3'].value == "TableB"

    ws_vba = wb["VBA_Diffs"]
    assert ws_vba['A1'].value == "Object"
    assert ws_vba['B1'].value == "Diff..."
    assert ws_vba['A2'].value == "Module1.bas"
    assert "-Old line" in ws_vba['B4'].value